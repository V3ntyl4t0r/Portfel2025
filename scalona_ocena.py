# scalona_ocena.py
import pandas as pd
import os
import subprocess
import walidacja_danych
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def score_company(row):
    score = 0

    # Fundamentalne
    if row["P/E"] != "N/A" and float(row["P/E"]) < 25:
        score += 2
    if row["ROE (%)"] != "N/A" and float(row["ROE (%)"]) > 15:
        score += 2
    if row["Debt/Assets"] != "N/A" and float(row["Debt/Assets"]) < 0.5:
        score += 1
    if row["PEG"] != "N/A" and 0.5 <= float(row["PEG"]) <= 1.5:
        score += 2
    if row["EV/FCF"] != "N/A" and float(row["EV/FCF"]) < 15:
        score += 2

    # Techniczne
    if row.get("RSI", "N/A") != "N/A":
        rsi = float(row["RSI"])
        if rsi < 30:
            score += 2
        elif rsi > 70:
            score -= 1
    if row.get("Drop from ATH (%)", "N/A") != "N/A" and float(row["Drop from ATH (%)"]) > 10:
        score += 1
    if row.get("SMA50", "N/A") != "N/A" and row.get("SMA200", "N/A") != "N/A":
        if float(row["SMA50"]) > float(row["SMA200"]):
            score += 2

    return score

def classify_company(score):
    if score >= 9:
        return "Dobra i tania"
    elif 6 <= score < 9:
        return "Dobra w dobrej cenie"
    elif 4 <= score < 6:
        return "Spółka średnia"
    else:
        return "Spółka słaba"

def classify_fundamental(row):
    if row["P/E"] != "N/A" and float(row["P/E"]) < 25 and \
       row["ROE (%)"] != "N/A" and float(row["ROE (%)"]) > 15 and \
       row["Debt/Assets"] != "N/A" and float(row["Debt/Assets"]) < 0.5:
        return "Strong"
    elif row["P/E"] != "N/A" or row["ROE (%)"] != "N/A" or row["Debt/Assets"] != "N/A":
        return "Medium"
    else:
        return "Weak"

def classify_valuation(row):
    try:
        rsi = float(row["RSI"]) if row.get("RSI", "N/A") != "N/A" else None
        drop = float(row["Drop from ATH (%)"]) if row.get("Drop from ATH (%)", "N/A") != "N/A" else None

        if rsi and rsi < 30 or (drop and drop > 20):
            return "Undervalued"
        elif rsi and rsi > 70:
            return "Overvalued"
        else:
            return "Fairly Valued"
    except:
        return "Unknown"

def add_legend_sheet(file_path):
    wb = load_workbook(file_path)
    if 'Legenda' in wb.sheetnames:
        del wb['Legenda']
    ws = wb.create_sheet("Legenda")

    content = [
        ["Legenda do scalona_ocena"], [],
        ["Fundamental Strength"],
        ["Strong: P/E < 25 AND ROE > 15% AND Debt/Assets < 0.5"],
        ["Medium: Częściowe spełnienie warunków"],
        ["Weak: Brak danych lub słabe wartości"], [],
        ["Valuation Status"],
        ["Undervalued: RSI < 30 lub Drop from ATH > 20%"],
        ["Overvalued: RSI > 70"],
        ["Fairly Valued: Pomiędzy undervalued a overvalued"],
        ["Unknown: Brak danych"], [],
        ["Ocena końcowa"],
        ["Dobra i tania: Score >= 9"],
        ["Dobra w dobrej cenie: Score 6-8"],
        ["Spółka średnia: Score 4-5"],
        ["Spółka słaba: Score < 4"], [],
        ["EMA Crossover"],
        ["EMA20 > EMA50: bullish crossover (sygnał kupna)"],
        ["EMA20 < EMA50: bearish crossover (sygnał sprzedaży)"], [],
        ["Strefa"],
        ["Swing High: Cena blisko 3-miesięcznego maksimum"],
        ["Swing Low: Cena blisko 3-miesięcznego minimum"],
        ["Neutral Zone: Cena w środku przedziału min-max"], [],
        ["Status OK"],
        ["OK: Brak alertów walidacji"],
        ["Inny komunikat: wykryta anomalia"]
    ]

    for row in content:
        ws.append(row)

    wb.save(file_path)

def format_excel_file(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    valuation_colors = {"Undervalued": "00FF00", "Fairly Valued": "FFFF00", "Overvalued": "FF0000", "Unknown": "FFFFFF"}
    fundamental_colors = {"Strong": "00FF00", "Medium": "FFFF00", "Weak": "FF0000"}
    ocena_colors = {"Dobra i tania": "008000", "Dobra w dobrej cenie": "00FF00", "Spółka średnia": "C0C0C0", "Spółka słaba": "FF0000"}

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter and cell.value:
                header = ws.cell(row=1, column=cell.column).value
                if header == "Valuation Status":
                    color = valuation_colors.get(cell.value, None)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if header == "Fundamental Strength":
                    color = fundamental_colors.get(cell.value, None)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                if header == "Ocena końcowa":
                    color = ocena_colors.get(cell.value, None)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    wb.save(file_path)

def merge_and_classify(fundamental_file, technical_file, output_file):
    df_fund = pd.read_csv(fundamental_file)
    df_tech = pd.read_csv(technical_file)

    df = pd.merge(df_fund, df_tech, on="Ticker", how="outer")
    df.drop_duplicates(subset=["Ticker"], inplace=True)

    df["Score"] = df.apply(score_company, axis=1)
    df["Ocena końcowa"] = df["Score"].apply(classify_company)
    df["Fundamental Strength"] = df.apply(classify_fundamental, axis=1)
    df["Valuation Status"] = df.apply(classify_valuation, axis=1)

    df = walidacja_danych.validate_data(df)
    df.sort_values(by="Score", ascending=False, inplace=True)

    cols = list(df.columns)
    for col in ["Score", "Ocena końcowa", "Fundamental Strength", "Valuation Status"]:
        cols.insert(2, cols.pop(cols.index(col)))

    # Przenieś RSI, Strefa, EMA Crossover za Price
    for col in ["EMA Crossover", "Strefa", "RSI"]:
        if col in cols:
            cols.insert(cols.index("Price") + 1, cols.pop(cols.index(col)))

    # Przenieś Date na koniec
    if "Date" in cols:
        cols.append(cols.pop(cols.index("Date")))

    df = df[cols]

    excel_file = output_file.replace(".csv", ".xlsx")
    df.to_excel(excel_file, index=False)

    format_excel_file(excel_file)
    add_legend_sheet(excel_file)

    print(f"Zapisano do: {excel_file}.")

    try:
        subprocess.run(["start", "excel", excel_file], shell=True)
    except:
        print("Nie udało się otworzyć pliku w Excelu.")

if __name__ == "__main__":
    fund_path = r"C:\Portfel2025\analiza_fundamentalna.csv"
    tech_path = r"C:\Portfel2025\analiza_techniczna.csv"
    out_path = r"C:\Portfel2025\scalona_ocena.csv"
    merge_and_classify(fund_path, tech_path, out_path)
