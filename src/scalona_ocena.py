# scalona_ocena.py
import pandas as pd
import subprocess
import walidacja_danych
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.ensemble import RandomForestClassifier
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import OneHotEncoder
from sklearn.impute import SimpleImputer
from sklearn.compose import ColumnTransformer
import numpy as np

def train_model(df, features, target_column="Target (6m +10%)"):
    df_model = df[df[target_column].isin([0, 1])]
    X = df_model[features]
    y = df_model[target_column]

    numeric = [f for f in features if f not in ["EMA Crossover", "Strefa"]]
    categorical = ["EMA Crossover", "Strefa"]

    # Usuń wiersze zawierające inf, -inf
    X = X.replace([np.inf, -np.inf], np.nan)
    X = X.dropna()
    y = y.loc[X.index]  # dopasuj y do X po odrzuceniu wierszy

    if len(X) == 0:
        print("❌ Brak danych do trenowania modelu – zbyt wiele braków.")
        df["ML_Predicted"] = 0
        df["ML_Points"] = 0
        df["Ocena AI"] = "Unikaj"
        return df, None

    preprocessor = ColumnTransformer([
        ("num", SimpleImputer(strategy="mean"), numeric),
        ("cat", OneHotEncoder(handle_unknown="ignore"), categorical)
    ])

    pipeline = Pipeline([
        ("preprocessor", preprocessor),
        ("model", RandomForestClassifier(n_estimators=100, random_state=42))
    ])

    pipeline.fit(X, y)
    df["ML_Predicted"] = pipeline.predict(df[features])

    # Feature importance
    classifier = pipeline.named_steps["model"]
    importances = classifier.feature_importances_
    transformed = pipeline.named_steps["preprocessor"].fit_transform(X)
    feature_names_num = numeric
    feature_names_cat = pipeline.named_steps["preprocessor"].transformers_[1][1].get_feature_names_out(categorical)
    feature_names = np.concatenate([feature_names_num, feature_names_cat])

    # Dopasuj długość tablic, by uniknąć błędu
    if len(importances) != len(feature_names):
        min_len = min(len(importances), len(feature_names))
        importances = importances[:min_len]
        feature_names = feature_names[:min_len]

    fi = pd.DataFrame({"Feature": feature_names, "Importance": importances})
    fi = fi.sort_values(by="Importance", ascending=False)
    fi.to_csv("../data/feature_importance.csv", index=False, encoding="utf-8-sig")
    print("\n📊 Top 10 cech wg ważności:")
    print(fi.head(10))

    df["ML_Points"] = pipeline.predict_proba(df[features])[:, 1] * 100

    def classify_ml_points(row):
        if row["ML_Points"] >= 80 and row["Score"] >= 6:
            return "Kupuj"
        elif row["ML_Points"] >= 60 and row["Score"] >= 5:
            return "Obserwuj"
        else:
            return "Unikaj"

    df["Ocena AI"] = df.apply(classify_ml_points, axis=1)

    return df, pipeline

def score_company(row):
    score = 0
    if row["P/E"] != "N/A" and float(row["P/E"]) < 25:
        score += 2
    if row["ROE (%)"] != "N/A" and float(row["ROE (%)"]) > 15:
        score += 2
    if row["Debt/Assets"] != "N/A" and float(row["Debt/Assets"]) < 0.5:
        score += 1
    if row["PEG"] != "N/A" and 0.5 <= float(row["PEG"]) <= 1.5:
        score += 2
    if row["EV/FCF"] != "N/A" and float(row["EV/FCF"]) < 15:
        score += 2
    if row.get("RSI", "N/A") != "N/A":
        rsi = float(row["RSI"])
        if rsi < 30:
            score += 2
        elif rsi > 70:
            score -= 1
    if row.get("Drop from ATH (%)", "N/A") != "N/A" and float(row["Drop from ATH (%)"]) > 10:
        score += 1
    if row.get("SMA50", "N/A") != "N/A" and row.get("SMA200", "N/A") != "N/A":
        if float(row["SMA50"]) > float(row["SMA200"]):
            score += 2
    return score

def classify_company(score):
    if score >= 9:
        return "Dobra i tania"
    elif 6 <= score < 9:
        return "Dobra w dobrej cenie"
    elif 4 <= score < 6:
        return "Spółka średnia"
    else:
        return "Spółka słaba"

def classify_fundamental(row):
    if row["P/E"] != "N/A" and float(row["P/E"]) < 25 and \
       row["ROE (%)"] != "N/A" and float(row["ROE (%)"]) > 15 and \
       row["Debt/Assets"] != "N/A" and float(row["Debt/Assets"]) < 0.5:
        return "Strong"
    elif row["P/E"] != "N/A" or row["ROE (%)"] != "N/A" or row["Debt/Assets"] != "N/A":
        return "Medium"
    else:
        return "Weak"

def classify_valuation(row):
    try:
        rsi = float(row["RSI"]) if row.get("RSI", "N/A") != "N/A" else None
        drop = float(row["Drop from ATH (%)"]) if row.get("Drop from ATH (%)", "N/A") != "N/A" else None
        if rsi and rsi < 30 or (drop and drop > 20):
            return "Undervalued"
        elif rsi and rsi > 70:
            return "Overvalued"
        else:
            return "Fairly Valued"
    except:
        return "Unknown"

def get_return(ticker, months=6):
    import yfinance as yf
    try:
        hist = yf.Ticker(ticker).history(period=f"{months+1}mo")
        if hist.empty or len(hist) < 2:
            return "N/A"
        p_now = hist["Close"].iloc[-1]
        p_then = hist["Close"].iloc[0]
        return 1 if ((p_now - p_then) / p_then) * 100 >= 10 else 0
    except:
        return "N/A"

def format_excel_file(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    colors = {
        "Undervalued": "00FF00", "Fairly Valued": "FFFF00", "Overvalued": "FF0000", "Unknown": "FFFFFF",
        "Strong": "00FF00", "Medium": "FFFF00", "Weak": "FF0000",
        "Dobra i tania": "008000", "Dobra w dobrej cenie": "00FF00", "Spółka średnia": "C0C0C0", "Spółka słaba": "FF0000",  "Kupuj": "00FF00", "Obserwuj": "FFFF00", "Unikaj": "FF0000"
    }

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header = ws.cell(row=1, column=cell.col_idx).value
            if header in ["Valuation Status", "Fundamental Strength", "Ocena końcowa"]:
                color = colors.get(cell.value)
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            if header == "ML_Predicted":
                cell.fill = PatternFill(start_color="00CCFF" if cell.value == 1 else "CCCCCC", fill_type="solid")

    wb.save(file_path)

def add_legend_sheet(file_path):
    wb = load_workbook(file_path)
    if 'Legenda' in wb.sheetnames:
        del wb['Legenda']
    ws = wb.create_sheet("Legenda")

    content = [
        ["Legenda"], [],
        ["ML_Predicted: 1 = Kupuj, 0 = Nie kupuj"], [],
        ["Valuation Status:"],
        ["Undervalued: RSI < 30 lub Drop from ATH > 20%"],
        ["Overvalued: RSI > 70"],
        ["Fairly Valued: Pomiędzy"],
        ["Unknown: Brak danych"], [],
        ["Fundamental Strength:"],
        ["Strong: P/E < 25 AND ROE > 15% AND Debt/Assets < 0.5"],
        ["Medium: Częściowo spełnione"],
        ["Weak: Brak danych lub słabe"], [],
        ["Ocena końcowa:"],
        ["Dobra i tania: Score >= 9"],
        ["Dobra w dobrej cenie: 6–8"],
        ["Średnia: 4–5"],
        ["Słaba: < 4"],
        ["Ocena AI:"],
        ["Kupuj: ML_Points >= 80 i Score >= 6"],
        ["Obserwuj: ML_Points >= 60 i Score >= 5"],
        ["Unikaj: pozostałe przypadki"],
    ]

    for row in content:
        ws.append(row)

    wb.save(file_path)

def merge_and_classify(fundamental_file, technical_file, output_file):
    df_fund = pd.read_csv(fundamental_file)
    df_tech = pd.read_csv(technical_file)
    df = pd.merge(df_fund, df_tech, on="Ticker", how="outer").drop_duplicates("Ticker")

    df["Score"] = df.apply(score_company, axis=1)
    df["Ocena końcowa"] = df["Score"].apply(classify_company)
    df["Fundamental Strength"] = df.apply(classify_fundamental, axis=1)
    df["Valuation Status"] = df.apply(classify_valuation, axis=1)
    df["Target (6m +10%)"] = df["Ticker"].apply(get_return)
    df = walidacja_danych.validate_data(df)

    features = [
        "P/E", "PEG", "ROE (%)", "Debt/Assets", "EV/FCF",
        "EPS Growth (%)", "Revenue Growth (%)", "RSI", "Drop from ATH (%)",
        "SMA50", "SMA200", "Beta", "Dividend Yield (%)",
        "EMA Crossover", "Strefa"
    ]
    df, _ = train_model(df, features)

    df.sort_values(by="Score", ascending=False, inplace=True)
    excel_path = output_file.replace(".csv", ".xlsx")
    df.to_excel(excel_path, index=False)
    format_excel_file(excel_path)
    add_legend_sheet(excel_path)

    print(f"Zapisano do: {excel_path}")
    try:
        subprocess.run(["start", "excel", excel_path], shell=True)
    except:
        pass

if __name__ == "__main__":
    merge_and_classify("../data/analiza_fundamentalna.csv", "../data/analiza_techniczna.csv", "../data/scalona_ocena.csv")
